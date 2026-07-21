"""XLSX sheet-level summaries, not raw cells."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook

from find_and_seek.ingest.extract.base import ExtractResult, TextBlock


def _sheet_summary(name: str, ws) -> str:
    headers: list[str] = []
    sample_rows: list[str] = []
    max_row = min(ws.max_row or 0, 50)
    max_col = min(ws.max_column or 0, 20)

    if max_row == 0 or max_col == 0:
        return f"Sheet '{name}': empty"

    for col in range(1, max_col + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            headers.append(str(val))

    for row in range(2, min(max_row + 1, 7)):
        cells = []
        for col in range(1, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                cells.append(str(val))
        if cells:
            sample_rows.append(", ".join(cells))

    header_str = ", ".join(headers) if headers else "no headers"
    sample_str = "; ".join(sample_rows[:3]) if sample_rows else "no data rows"
    return (
        f"Sheet '{name}': columns [{header_str}]. "
        f"Sample rows: {sample_str}. "
        f"Dimensions: ~{max_row} rows x {max_col} cols."
    )


def extract_xlsx(path: Path) -> ExtractResult:
    # Read into BytesIO rather than passing the path string: openpyxl
    # validates the FILENAME's suffix against a fixed allowlist before ever
    # looking at content, so a real xlsx reached via router.py's magic-byte
    # sniff (e.g. a workbook saved under a .txt/.md/.csv suffix) is rejected
    # on the extension alone. A file object carries no filename, so that
    # check never fires.
    with open(path, "rb") as f:
        wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
    blocks: list[TextBlock] = []
    for idx, name in enumerate(wb.sheetnames, start=1):
        ws = wb[name]
        summary = _sheet_summary(name, ws)
        blocks.append(
            TextBlock(
                text=summary,
                source_type="table",
                location_ref=f"sheet {idx}: {name}",
            )
        )
    wb.close()
    return ExtractResult(blocks=blocks, file_type="spreadsheet")
